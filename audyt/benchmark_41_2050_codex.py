# -*- coding: utf-8 -*-
"""Benchmark V3 na zleceniu 41_2050.

Skrypt celowo NIE czyta folderu `sprawdzenie`. Wejscia sa jawnie ograniczone do:
  - DOK41_CONV
  - DXF_do_gotowe
  - wykaz APROBOCZY2_41.45672050_lista*.xlsx

Wyniki trafia do `sprawdzenie_codex`.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path

import ezdxf
import matplotlib.pyplot as plt
import openpyxl
from ezdxf.addons.drawing import Frontend, RenderContext
from ezdxf.addons.drawing.matplotlib import MatplotlibBackend
from openpyxl.styles import Alignment, Font, PatternFill


REPO = Path(__file__).resolve().parents[1]
BASE = Path(r"\\QNAP-ENERGO\tmp\EKSTRAKTOR_DXF_test\41_2050_test")
DOK = BASE / "DOK41_CONV"
WZORCE = BASE / "DXF_do_gotowe"
OUT = BASE / "sprawdzenie_codex"
PROD = REPO / "produkcja"

sys.path.insert(0, str(PROD))
import ocena  # noqa: E402


TARGET_STATUSES = {"RYSUJ", "Z BAZY"}
TOL_MM = 1.0
TOL_PROC = 0.2
RESULT_COLS = [
    "plik_dxf",
    "wymiar_gen",
    "wymiar_vs_wykaz",
    "interior_gen",
    "interior_wzorzec",
    "okregi_gen",
    "okregi_wzorzec",
    "kontur_otwarty",
    "zgodne_ze_wzorcem",
    "rozjazd",
    "ANALIZA",
    "DO_PRZEGLADU",
    "UWAGI_OPERATORA",
]


def norm_zeinr(v) -> str:
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if s.isdigit():
        return str(int(s))
    return s


def parse_dim(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def safe_stem(name: str) -> str:
    return str(name).strip().removesuffix(".dxf").removesuffix(".DXF")


def test_filename(nazwa: str) -> str:
    stem = safe_stem(nazwa).rstrip("_")
    return f"{stem}_test.dxf"


def read_targets(wykaz_path: Path):
    wb = openpyxl.load_workbook(wykaz_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    headers = {
        str(ws.cell(1, c).value).strip().lower(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }
    required = ["sprawdzenie", "nazwa", "zeinr", "posn", "zakupy", "abmess_1", "abmes_2"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise RuntimeError(f"Brak kolumn w wykazie: {missing}")

    targets = []
    for row in range(2, ws.max_row + 1):
        spraw = ws.cell(row, headers["sprawdzenie"]).value
        if not isinstance(spraw, str) or spraw.strip().upper() not in TARGET_STATUSES:
            continue
        z = norm_zeinr(ws.cell(row, headers["zeinr"]).value)
        posn = int(ws.cell(row, headers["posn"]).value)
        nazwa = safe_stem(ws.cell(row, headers["nazwa"]).value)
        d1 = parse_dim(ws.cell(row, headers["abmess_1"]).value)
        d2 = parse_dim(ws.cell(row, headers["abmes_2"]).value)
        targets.append(
            {
                "row": row,
                "zeinr": z,
                "posn": posn,
                "nazwa": nazwa,
                "sprawdzenie": spraw.strip().upper(),
                "zakupy": str(ws.cell(row, headers["zakupy"]).value).strip(),
                "wykaz_dim": (max(d1, d2), min(d1, d2)) if d1 and d2 else None,
            }
        )
    return targets


def find_wykaz() -> Path:
    files = list(BASE.glob("APROBOCZY2_41.45672050_lista*.xlsx"))
    if not files:
        raise FileNotFoundError("Nie znaleziono wykazu APROBOCZY2_41.45672050_lista*.xlsx")
    return files[0]


def build_source_index():
    by_key = defaultdict(list)
    for p in DOK.glob("*"):
        if p.suffix.lower() not in {".dxf", ".dwg"}:
            continue
        stem = p.stem
        keys = {norm_zeinr(stem)}
        if stem.endswith("_1"):
            keys.add(norm_zeinr(stem[:-2]))
        if stem.endswith("_conv"):
            keys.add(norm_zeinr(stem[:-5]))
        if stem.endswith("_1_conv"):
            keys.add(norm_zeinr(stem[:-7]))
        for k in keys:
            by_key[k].append(p)
    return by_key


def choose_source(zeinr: str, source_index, conv_dir: Path):
    candidates = source_index.get(norm_zeinr(zeinr), [])
    dxfs = [p for p in candidates if p.suffix.lower() == ".dxf"]
    dwgs = [p for p in candidates if p.suffix.lower() == ".dwg"]

    def score(p: Path):
        s = p.stem
        exact = s == zeinr
        exact_1 = s == f"{zeinr}_1"
        return (0 if exact else 1 if exact_1 else 2, len(s), s)

    if dxfs:
        return sorted(dxfs, key=score)[0], "dxf"
    if not dwgs:
        return None, "brak"
    src = sorted(dwgs, key=score)[0]
    dst = conv_dir / f"{src.stem}_conv.dxf"
    if dst.exists():
        return dst, "dwg_juz_skonwertowany"
    cmd = [sys.executable, str(PROD / "silniki" / "convert_dwg.py"), str(src), str(dst)]
    result = subprocess.run(cmd, cwd=REPO, capture_output=True, text=True, timeout=180)
    if result.returncode != 0 or not dst.exists():
        raise RuntimeError(
            f"Konwersja DWG nieudana dla {src.name}: rc={result.returncode}\n"
            f"STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
        )
    return dst, "dwg_skonwertowany"


_WZ_RE = re.compile(r"(SL\d+|\d+)_p(\d+)", re.IGNORECASE)


def build_wzorzec_index():
    exact = {}
    by_key = defaultdict(list)
    for p in WZORCE.rglob("*.dxf"):
        exact[p.stem.lower()] = p
        m = _WZ_RE.search(p.stem)
        if m:
            by_key[(norm_zeinr(m.group(1).upper()), int(m.group(2)))].append(p)
    return exact, by_key


def choose_wzorzec(row, exact, by_key):
    stem = row["nazwa"].lower()
    if stem in exact:
        return exact[stem], "exact"
    candidates = by_key.get((row["zeinr"], row["posn"]), [])
    if len(candidates) == 1:
        return candidates[0], "zeinr_posn"
    if candidates:
        # Prefer matching thickness/material prefix if exact formula suffix differs.
        prefix = row["nazwa"].split(f"_{row['zeinr']}_p{row['posn']}", 1)[0].lower()
        pref = [p for p in candidates if p.stem.lower().startswith(prefix)]
        if len(pref) == 1:
            return pref[0], "prefix"
        return sorted(candidates, key=lambda p: p.name.lower())[0], "ambiguous_first"
    return None, "brak"


def run_v3_for_sources(targets, force=False, limit=None, start_index=1, batch_size=None):
    out_raw = OUT / "V3_raw"
    logs = OUT / "logs"
    conv_dir = OUT / "konwersje"
    out_raw.mkdir(parents=True, exist_ok=True)
    logs.mkdir(parents=True, exist_ok=True)
    conv_dir.mkdir(parents=True, exist_ok=True)

    wykaz_path = find_wykaz()
    source_index = build_source_index()
    zeinrs = sorted({t["zeinr"] for t in targets})
    if start_index > 1:
        zeinrs = zeinrs[start_index - 1 :]
    if batch_size:
        zeinrs = zeinrs[:batch_size]
    if limit:
        zeinrs = zeinrs[:limit]

    summary = {}
    for i, z in enumerate(zeinrs, 1):
        raw_dir = out_raw / z
        ocena_csv = raw_dir / f"{z}_ocena.csv"
        if ocena_csv.exists() and not force:
            summary[z] = {"status": "POMINIETO_EXISTING", "raw_dir": str(raw_dir)}
            print(f"[{i}/{len(zeinrs)}] {z}: istnieje ocena.csv - pomijam")
            continue
        try:
            src, src_status = choose_source(z, source_index, conv_dir)
            if src is None:
                summary[z] = {"status": "BRAK_ZRODLA", "raw_dir": str(raw_dir)}
                print(f"[{i}/{len(zeinrs)}] {z}: BRAK ZRODLA")
                continue
            raw_dir.mkdir(parents=True, exist_ok=True)
            cmd = [sys.executable, str(PROD / "orkiestrator.py"), str(src), str(wykaz_path), str(raw_dir)]
            print(f"[{i}/{len(zeinrs)}] {z}: V3 z {src.name} ({src_status})")
            t0 = time.time()
            result = subprocess.run(cmd, cwd=REPO, capture_output=True, text=True, timeout=420)
            elapsed = round(time.time() - t0, 1)
            log_text = (
                f"CMD: {' '.join(cmd)}\nSRC_STATUS: {src_status}\nRC: {result.returncode}\n"
                f"ELAPSED_S: {elapsed}\n\nSTDOUT:\n{result.stdout}\n\nSTDERR:\n{result.stderr}\n"
            )
            (logs / f"{z}.log").write_text(log_text, encoding="utf-8", errors="replace")
            status = "OK" if result.returncode == 0 else f"RC_{result.returncode}"
            if result.returncode != 0:
                print(f"  GLOSNO: {z} zakonczyl sie {status}")
            summary[z] = {
                "status": status,
                "source": str(src),
                "source_status": src_status,
                "raw_dir": str(raw_dir),
                "elapsed_s": elapsed,
            }
        except Exception as exc:
            summary[z] = {"status": "BLAD", "error": str(exc), "raw_dir": str(raw_dir)}
            (logs / f"{z}.error.txt").write_text(str(exc), encoding="utf-8", errors="replace")
            print(f"  GLOSNO: {z}: {exc}")

    status_path = OUT / "generowanie_status.json"
    if status_path.exists():
        try:
            previous = json.loads(status_path.read_text(encoding="utf-8"))
            previous.update(summary)
            summary = previous
        except Exception:
            pass
    status_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def read_ocena_row(raw_dir: Path, zeinr: str, posn: int):
    path = raw_dir / f"{zeinr}_ocena.csv"
    if not path.exists():
        return None
    with open(path, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f, delimiter=";"):
            try:
                if int(r.get("posn", -1)) == int(posn):
                    return r
            except ValueError:
                continue
    return None


def raw_generated_path(raw_dir: Path, zeinr: str, posn: int, ocena_row):
    if ocena_row:
        name = ocena_row.get("plik_produkcyjny")
        if name and name != "-":
            p = raw_dir / name
            if p.exists():
                return p
    candidates = sorted(raw_dir.glob(f"{zeinr}_p{posn}*.dxf"))
    return candidates[0] if candidates else None


def measure(path: Path, label: str):
    if path is None or not path.exists():
        return None, [f"brak pliku {label}"]
    try:
        return ocena.zmierz_wariant({"nazwa": label, "dxf_path": str(path)}), []
    except Exception as exc:
        return None, [f"blad pomiaru {label}: {exc}"]


def fmt_dim(m):
    if not m:
        return ""
    return f"{m['wymiar_x']:.2f} x {m['wymiar_y']:.2f}"


def as_int_or_none(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def classify(row, gen_m, wz_m, gen_vs_wykaz, zgodne, roz, zrodlo_prawda):
    reasons = list(roz)
    review = False
    analysis = "OK"
    if gen_m is None:
        return "BRAK_GENERATU", "TAK", "brak wygenerowanego DXF"
    if wz_m is None:
        return "BRAK_WZORCA", "TAK", "brak wzorca do porownania"

    if gen_vs_wykaz[0] is False:
        review = True
        reasons.append(f"wymiar vs wykaz: {gen_vs_wykaz[1]}")
    if gen_m["niedomkniete"]:
        review = True
        reasons.append(f"GEN otwarty kontur/brud={gen_m['brud']}")
    if not zgodne:
        review = True
    if gen_m["interior"] != wz_m["interior"]:
        review = True
    if gen_m["circles_dedup"] != wz_m["circles_dedup"]:
        review = True
    if wz_m["niedomkniete"] != gen_m["niedomkniete"]:
        review = True

    better = []
    src_truth = as_int_or_none(zrodlo_prawda)
    if src_truth is not None and gen_m["interior"] == src_truth and wz_m["interior"] != src_truth:
        better.append(
            f"interior GEN={gen_m['interior']} zgodny ze zrodlo_prawda={src_truth}, "
            f"wzorzec={wz_m['interior']}"
        )
    if not gen_m["niedomkniete"] and wz_m["niedomkniete"]:
        better.append(f"GEN domkniety, wzorzec otwarty/brud={wz_m['brud']}")
    wz_vs_wykaz = ocena.wymiar_zgodny(
        wz_m["wymiar_x"], wz_m["wymiar_y"], row["wykaz_dim"], TOL_MM, TOL_PROC
    )
    if gen_vs_wykaz[0] is True and wz_vs_wykaz[0] is False:
        better.append(f"GEN wymiar OK vs wykaz, wzorzec: {wz_vs_wykaz[1]}")
    if src_truth is not None and gen_m["interior"] == src_truth and gen_m["circles_dedup"] > wz_m["circles_dedup"]:
        better.append(
            f"mozliwa degradacja otworow we wzorcu: okregi GEN={gen_m['circles_dedup']} "
            f"wzorzec={wz_m['circles_dedup']}"
        )

    if better:
        analysis = "V3_LEPSZY"
        review = True
        reasons = [f"V3_LEPSZY: {b}" for b in better] + reasons
    elif review:
        analysis = "ROZJAZD"

    return analysis, "TAK" if review else "NIE", "; ".join(dict.fromkeys(reasons))


def render_pair(gen_path: Path | None, wz_path: Path | None, out_png: Path, title: str):
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig, axes = plt.subplots(1, 2, figsize=(16, 7))
    fig.patch.set_facecolor("black")
    labels = [("V3", gen_path), ("WZORZEC", wz_path)]
    for ax, (label, path) in zip(axes, labels):
        ax.set_facecolor("black")
        ax.set_aspect("equal", adjustable="datalim")
        ax.axis("off")
        ax.set_title(label, color="white", fontsize=11)
        if path is None or not path.exists():
            ax.text(0.5, 0.5, "BRAK DXF", color="red", ha="center", va="center", transform=ax.transAxes)
            continue
        try:
            doc = ezdxf.readfile(str(path))
            Frontend(RenderContext(doc), MatplotlibBackend(ax)).draw_layout(doc.modelspace())
            ax.autoscale()
        except Exception as exc:
            ax.text(
                0.5,
                0.5,
                f"BLAD RENDERU\n{exc}",
                color="red",
                ha="center",
                va="center",
                transform=ax.transAxes,
            )
    fig.suptitle(title, color="white", fontsize=12)
    fig.tight_layout()
    fig.savefig(out_png, dpi=140, facecolor="black")
    plt.close(fig)


def compare_and_copy(targets):
    raw_root = OUT / "V3_raw"
    gen_dir = OUT / "DXF_wygenerowane"
    wz_dir = OUT / "DXF_wzorce"
    png_dir = OUT / "PNG_porownanie"
    gen_dir.mkdir(parents=True, exist_ok=True)
    wz_dir.mkdir(parents=True, exist_ok=True)
    png_dir.mkdir(parents=True, exist_ok=True)

    exact, by_key = build_wzorzec_index()
    rows = []
    for idx, row in enumerate(targets, 1):
        zeinr = row["zeinr"]
        posn = row["posn"]
        raw_dir = raw_root / zeinr
        oc_row = read_ocena_row(raw_dir, zeinr, posn)
        zrodlo_prawda = oc_row.get("zrodlo_prawda") if oc_row else None
        gen_raw = raw_generated_path(raw_dir, zeinr, posn, oc_row)
        gen_out = gen_dir / test_filename(row["nazwa"])
        if gen_raw and gen_raw.exists():
            shutil.copy2(gen_raw, gen_out)
        else:
            gen_out = None

        wz, wz_match = choose_wzorzec(row, exact, by_key)
        wz_copy = None
        if wz and wz.exists():
            wz_copy = wz_dir / wz.name
            if not wz_copy.exists():
                shutil.copy2(wz, wz_copy)

        gen_m, gen_err = measure(gen_out, "GEN")
        wz_m, wz_err = measure(wz_copy, "WZORZEC")

        roz = []
        zgodne = False
        if gen_m and wz_m:
            zgodne, rz = ocena.sygnatury_zgodne(gen_m, wz_m, TOL_MM, TOL_PROC)
            roz.extend(rz)
        roz.extend(gen_err)
        roz.extend(wz_err)
        gen_vs_wykaz = (None, "brak generatu")
        if gen_m:
            gen_vs_wykaz = ocena.wymiar_zgodny(
                gen_m["wymiar_x"], gen_m["wymiar_y"], row["wykaz_dim"], TOL_MM, TOL_PROC
            )
        analiza, do_przegladu, rozjazd = classify(
            row, gen_m, wz_m, gen_vs_wykaz, zgodne, roz, zrodlo_prawda
        )

        png_path = png_dir / (gen_out.stem + ".png" if gen_out else f"{row['nazwa']}_BRAK_GEN.png")
        render_pair(gen_out, wz_copy, png_path, f"{idx}/75 row {row['row']} {zeinr} p{posn}")

        rows.append(
            {
                **row,
                "plik_dxf": gen_out.name if gen_out else "",
                "wzorzec": wz_copy.name if wz_copy else "",
                "wzorzec_match": wz_match,
                "png": png_path.name,
                "wymiar_gen": fmt_dim(gen_m),
                "wymiar_vs_wykaz": "ok" if gen_vs_wykaz[0] is True else gen_vs_wykaz[1],
                "interior_gen": "" if gen_m is None else gen_m["interior"],
                "interior_wzorzec": "" if wz_m is None else wz_m["interior"],
                "okregi_gen": "" if gen_m is None else gen_m["circles_dedup"],
                "okregi_wzorzec": "" if wz_m is None else wz_m["circles_dedup"],
                "kontur_otwarty": (
                    ""
                    if gen_m is None and wz_m is None
                    else f"gen={bool(gen_m and gen_m['niedomkniete'])}; "
                    f"wzorzec={bool(wz_m and wz_m['niedomkniete'])}"
                ),
                "zgodne_ze_wzorcem": "TAK" if zgodne else "NIE",
                "rozjazd": rozjazd,
                "ANALIZA": analiza,
                "DO_PRZEGLADU": do_przegladu,
                "UWAGI_OPERATORA": "",
                "zrodlo_prawda": zrodlo_prawda,
                "v3_semafor": oc_row.get("semafor") if oc_row else "",
                "v3_powod": oc_row.get("powod") if oc_row else "",
            }
        )
        print(f"[POROWNANIE {idx}/{len(targets)}] row {row['row']} {zeinr} p{posn}: {analiza}")
    return rows


def write_csv(rows):
    out_csv = OUT / "benchmark_41_2050_codex.csv"
    fields = [
        "row",
        "zeinr",
        "posn",
        "nazwa",
        "sprawdzenie",
        "zakupy",
        "plik_dxf",
        "wzorzec",
        "wzorzec_match",
        "png",
        "wymiar_gen",
        "wymiar_vs_wykaz",
        "interior_gen",
        "interior_wzorzec",
        "okregi_gen",
        "okregi_wzorzec",
        "kontur_otwarty",
        "zgodne_ze_wzorcem",
        "rozjazd",
        "ANALIZA",
        "DO_PRZEGLADU",
        "UWAGI_OPERATORA",
        "zrodlo_prawda",
        "v3_semafor",
        "v3_powod",
    ]
    with open(out_csv, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter=";", extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    return out_csv


def write_workbook(wykaz_path: Path, rows):
    out_xlsx = OUT / "APROBOCZY2_41_2050_wykaz_ANALIZA_CODEX.xlsx"
    shutil.copy2(wykaz_path, out_xlsx)
    wb = openpyxl.load_workbook(out_xlsx, data_only=False)
    ws = wb.worksheets[0]
    start_col = ws.max_column + 1
    col_map = {}
    for offset, name in enumerate(RESULT_COLS):
        c = start_col + offset
        col_map[name] = c
        cell = ws.cell(1, c, name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    by_row = {r["row"]: r for r in rows}
    fill_review = PatternFill("solid", fgColor="FFF2CC")
    fill_bad = PatternFill("solid", fgColor="F4CCCC")
    fill_better = PatternFill("solid", fgColor="D9EAD3")
    for sheet_row, data in by_row.items():
        for name in RESULT_COLS:
            ws.cell(sheet_row, col_map[name], data.get(name, ""))
            ws.cell(sheet_row, col_map[name]).alignment = Alignment(wrap_text=True, vertical="top")
        fill = None
        if data["ANALIZA"] == "V3_LEPSZY":
            fill = fill_better
        elif data["DO_PRZEGLADU"] == "TAK":
            fill = fill_bad if data["ANALIZA"].startswith("BRAK") else fill_review
        if fill:
            for c in range(start_col, start_col + len(RESULT_COLS)):
                ws.cell(sheet_row, c).fill = fill

    widths = {
        "plik_dxf": 42,
        "wymiar_gen": 18,
        "wymiar_vs_wykaz": 30,
        "interior_gen": 14,
        "interior_wzorzec": 16,
        "okregi_gen": 12,
        "okregi_wzorzec": 15,
        "kontur_otwarty": 25,
        "zgodne_ze_wzorcem": 18,
        "rozjazd": 80,
        "ANALIZA": 18,
        "DO_PRZEGLADU": 16,
        "UWAGI_OPERATORA": 35,
    }
    for name, c in col_map.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = widths.get(name, 18)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = ws.freeze_panes or "A2"
    wb.save(out_xlsx)
    return out_xlsx


def write_report(rows, out_csv: Path, out_xlsx: Path):
    cnt = Counter(r["ANALIZA"] for r in rows)
    review = sum(1 for r in rows if r["DO_PRZEGLADU"] == "TAK")
    missing_gen = [r for r in rows if r["ANALIZA"] == "BRAK_GENERATU"]
    missing_wz = [r for r in rows if r["ANALIZA"] == "BRAK_WZORCA"]
    top_roz = [r for r in rows if r["DO_PRZEGLADU"] == "TAK"][:25]
    lines = [
        "# Benchmark V3 - 41_2050 (Codex)",
        "",
        f"Data uruchomienia: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"Zakres: {len(rows)} pozycji (sprawdzenie=RYSUJ lub Z BAZY).",
        "",
        "## Wynik liczbowy",
        "",
        f"- OK: {cnt.get('OK', 0)}",
        f"- ROZJAZD: {cnt.get('ROZJAZD', 0)}",
        f"- V3_LEPSZY: {cnt.get('V3_LEPSZY', 0)}",
        f"- BRAK_GENERATU: {cnt.get('BRAK_GENERATU', 0)}",
        f"- BRAK_WZORCA: {cnt.get('BRAK_WZORCA', 0)}",
        f"- DO_PRZEGLADU: {review}",
        "",
        "## Pliki",
        "",
        f"- Wykaz analiza: `{out_xlsx}`",
        f"- CSV analiza: `{out_csv}`",
        f"- DXF wygenerowane: `{OUT / 'DXF_wygenerowane'}`",
        f"- PNG porownanie: `{OUT / 'PNG_porownanie'}`",
        f"- Logi V3: `{OUT / 'logs'}`",
        "",
        "## Pozycje do przegladu (pierwsze 25)",
        "",
    ]
    for r in top_roz:
        lines.append(
            f"- row {r['row']} {r['zeinr']} p{r['posn']} {r['ANALIZA']}: {r['rozjazd'][:300]}"
        )
    if missing_gen:
        lines += ["", "## Brak generatu", ""]
        for r in missing_gen:
            lines.append(f"- row {r['row']} {r['zeinr']} p{r['posn']} {r['nazwa']}")
    if missing_wz:
        lines += ["", "## Brak wzorca", ""]
        for r in missing_wz:
            lines.append(f"- row {r['row']} {r['zeinr']} p{r['posn']} {r['nazwa']}")
    report = OUT / "raport_41_2050_codex.md"
    report.write_text("\n".join(lines), encoding="utf-8")
    return report


def verify_outputs(rows, out_xlsx: Path):
    gen_count = len(list((OUT / "DXF_wygenerowane").glob("*.dxf")))
    png_count = len(list((OUT / "PNG_porownanie").glob("*.png")))
    wb = openpyxl.load_workbook(out_xlsx, data_only=False, read_only=True)
    ws = wb.worksheets[0]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    missing_cols = [c for c in RESULT_COLS if c not in headers]
    return {
        "rows": len(rows),
        "generated_dxf": gen_count,
        "png": png_count,
        "xlsx_exists": out_xlsx.exists(),
        "missing_result_cols": missing_cols,
    }


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("--skip-generate", action="store_true")
    ap.add_argument("--only-generate", action="store_true")
    ap.add_argument("--map-only", action="store_true")
    ap.add_argument("--force-generate", action="store_true")
    ap.add_argument("--limit-generate", type=int, default=None)
    ap.add_argument("--start-index", type=int, default=1)
    ap.add_argument("--batch-size", type=int, default=None)
    args = ap.parse_args(argv)

    OUT.mkdir(parents=True, exist_ok=True)
    wykaz_path = find_wykaz()
    targets = read_targets(wykaz_path)
    (OUT / "target_rows.json").write_text(
        json.dumps(targets, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    print(f"Target rows: {len(targets)}")
    if len(targets) != 75:
        print(f"GLOSNO: oczekiwano 75 pozycji, jest {len(targets)}")

    if args.map_only:
        source_index = build_source_index()
        exact, by_key = build_wzorzec_index()
        missing_sources = []
        missing_wzorce = []
        for row in targets:
            if not source_index.get(row["zeinr"]):
                missing_sources.append((row["row"], row["zeinr"], row["posn"]))
            wz, _ = choose_wzorzec(row, exact, by_key)
            if wz is None:
                missing_wzorce.append((row["row"], row["zeinr"], row["posn"], row["nazwa"]))
        print(f"Unikalne rysunki: {len({t['zeinr'] for t in targets})}")
        print(f"Brak zrodel: {missing_sources}")
        print(f"Brak wzorcow: {missing_wzorce}")
        return 0

    if not args.skip_generate:
        run_v3_for_sources(
            targets,
            force=args.force_generate,
            limit=args.limit_generate,
            start_index=args.start_index,
            batch_size=args.batch_size,
        )
    if args.only_generate:
        return 0

    rows = compare_and_copy(targets)
    out_csv = write_csv(rows)
    out_xlsx = write_workbook(wykaz_path, rows)
    report = write_report(rows, out_csv, out_xlsx)
    verification = verify_outputs(rows, out_xlsx)
    (OUT / "verification.json").write_text(
        json.dumps(verification, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(verification, ensure_ascii=False, indent=2))
    print(f"CSV: {out_csv}")
    print(f"XLSX: {out_xlsx}")
    print(f"RAPORT: {report}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
