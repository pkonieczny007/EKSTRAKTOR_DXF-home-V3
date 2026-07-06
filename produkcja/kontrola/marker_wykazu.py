# -*- coding: utf-8 -*-
"""
MARKER WYKAZU (standard 2026-07-07) - zaznaczanie wynikow ekstrakcji w KOPII wykazu
dla zlecenia WIELORYSUNKOWEGO, w stylu operatora (pl.PLIK_DO_SPRAWDZENIA):

  X-DXF, Y-DXF (pomiar DXF int) + Abmess (numerycznie) + FORMULY 'x =x,y' / 'y =x,y'
  (=OR(...+-1)) w Excelu -> UWAGI = "ok" gdy obie PRAWDA, inaczej "sprawdz" (rece operatora).

Zasady: (1) kolumny wynikowe dopisywane ZA OSTATNIA UZYWANA kolumna (ws.max_column+1),
NIE za ostatnim nazwanym naglowkiem - inaczej nachodza na formuly budujace NAZWA
(bug 2026-07-07). (2) Oryginal NIETKNIETY - piszemy do KOPII. (3) Wykaz = BAZA
([[wykaz-to-baza-additive]]). (4) Per-zeinr scoping raportow (posn kolidowalby miedzy
rysunkami). (5) Formuly zapisane TAKZE jako tekst - operator dostraja recznie.

Po zmianie: python testy/test_marker_wykazu.py.
"""
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path

import ezdxf
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(HERE.parent))          # produkcja/
sys.path.insert(0, str(HERE.parent / "silniki"))
import raport  # noqa: E402
import fazowanie  # noqa: E402
import kontrola_wymiaru as kw  # noqa: E402
from extract_positions import parse_dim  # noqa: E402

# kolejnosc kolumn wynikowych (dopisywane na KONCU). status barwiony semaforem.
KOL = ["status", "plik_dxf", "technologia", "skala", "X-DXF", "Y-DXF",
       "Abmess_1", "Abmes_2", "x =x,y", "y =x,y", "UWAGI", "powod"]


def _zbierz_rekordy(all_folder, conv_dir, conv_pat):
    """{(zeinr,posn): rekord} z ocen wszystkich rysunkow (per-zeinr scoping)."""
    all_folder = Path(all_folder)
    conv_dir = Path(conv_dir)
    out = {}
    for ocena in sorted(all_folder.glob("*_ocena.csv")):
        zeinr = ocena.name.replace("_ocena.csv", "")
        raporty = {}
        for pod in ("warianty/wA", "warianty/wB"):
            rp = all_folder / pod / f"{zeinr}_raport.csv"
            if rp.exists():
                for r in raport._czytaj_csv(rp):
                    try:
                        raporty.setdefault(int(r["posn"]), r)
                    except (KeyError, ValueError):
                        pass
        conv = conv_dir / conv_pat.format(z=zeinr)
        ma_spaw = None
        if conv.exists():
            try:
                ma_spaw = raport.has_weld(ezdxf.readfile(str(conv)).modelspace())
            except Exception:
                pass
        for b in raport._czytaj_csv(ocena):
            try:
                posn = int(b["posn"])
            except (KeyError, ValueError):
                continue
            plik = b.get("plik_produkcyjny", "-")
            rap = raporty.get(posn)
            technologia = ""
            if ma_spaw is not None and rap is not None:
                try:
                    mg = int(float(rap.get("n_bend", 0) or 0)) > 0
                    technologia = raport.technologia_pozycji(ma_spaw, mg)
                except (ValueError, TypeError):
                    pass
            dxf_path = all_folder / plik if plik and plik != "-" else None
            if dxf_path and dxf_path.exists():
                try:
                    nf, fk = fazowanie.oznacz_w_pliku(dxf_path)
                    if nf:
                        technologia = (f"{technologia}+faza"
                                       if technologia and technologia != "brak" else "faza")
                except Exception:
                    pass
            out[(zeinr, posn)] = dict(
                semafor=raport.norm_semafor(b.get("semafor")),
                plik_dxf=plik or "-",
                technologia=technologia,
                powod=b.get("powod", ""),
                dxf_int=kw.zmierz_dxf_int(dxf_path) if dxf_path else None)
    return out


def zaznacz(all_folder, wykaz, out_path, conv_dir, conv_pat="{z}_1_conv.dxf",
            filtr_kolumna="sprawdzenie", filtr_pred=lambda s: "rysuj" in s):
    """Zaznacza wynik w KOPII wykazu. filtr_pred(str_lower)->bool: ktore wiersze.
    Zwraca (out_path, statystyki) albo (None, {}). GLOSNO przy bledach."""
    wykaz, out_path = Path(wykaz), Path(out_path)
    if not wykaz.exists():
        print(f"[MARKER] wykaz nie istnieje: {wykaz} (GLOSNO)")
        return None, {}
    rek = _zbierz_rekordy(all_folder, conv_dir, conv_pat)

    shutil.copy(wykaz, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.worksheets[0]
    hr, nag = raport._znajdz_naglowek(ws)
    if hr is None:
        print(f"[MARKER] brak naglowka Zeinr+Posn w {wykaz.name} (GLOSNO)")
        return None, {}
    # FIX: za ostatnia UZYWANA kolumna, nie za ostatnim nazwanym naglowkiem
    kol = dict(nag)
    nast = ws.max_column + 1
    for nazwa in KOL:
        if nazwa in kol:            # nie nadpisuj istniejacej kolumny wykazu
            nazwa_out = nazwa + "_v3"
        else:
            nazwa_out = nazwa
        ws.cell(row=hr, column=nast, value=nazwa_out)
        kol[nazwa] = nast
        nast += 1
    L = {k: get_column_letter(v) for k, v in kol.items()}
    ci_abm1_src = nag.get("Abmess_1")
    ci_abm2_src = nag.get("Abmes_2")

    sem = Counter()
    zap = 0
    filtr_col = nag.get(filtr_kolumna)
    for i in range(hr + 1, ws.max_row + 1):
        if filtr_col is not None:
            fv = ws.cell(row=i, column=filtr_col).value
            if not filtr_pred(str(fv).lower() if fv else ""):
                continue
        zc = ws.cell(row=i, column=kol["Zeinr"]).value
        pc = ws.cell(row=i, column=kol["Posn"]).value
        if zc is None or pc is None:
            continue
        try:
            posn = int(pc)
        except (ValueError, TypeError):
            continue
        # Abmess z wykazu (BAZA) - parsuj do liczb dla formuly
        a1 = parse_dim(ws.cell(row=i, column=ci_abm1_src).value) if ci_abm1_src else None
        a2 = parse_dim(ws.cell(row=i, column=ci_abm2_src).value) if ci_abm2_src else None
        # dopasuj rekord po (zeinr,posn) (substring zeinr jak raport)
        r = None
        for (zk, pk), rr in rek.items():
            if pk == posn and zk in str(zc):
                r = rr
                break
        dxf_int = r["dxf_int"] if r else None
        chk = kw.sprawdz(dxf_int, a1, a2)

        ws.cell(row=i, column=kol["status"], value=(r["semafor"] if r else "brak wyniku"))
        ws.cell(row=i, column=kol["plik_dxf"], value=(r["plik_dxf"] if r else "-"))
        ws.cell(row=i, column=kol["technologia"], value=(r["technologia"] if r else ""))
        ws.cell(row=i, column=kol["skala"], value=kw.skala(dxf_int, a1, a2))
        ws.cell(row=i, column=kol["X-DXF"], value=chk["x_dxf"])
        ws.cell(row=i, column=kol["Y-DXF"], value=chk["y_dxf"])
        ws.cell(row=i, column=kol["Abmess_1"], value=(round(a1) if a1 else ""))
        ws.cell(row=i, column=kol["Abmes_2"], value=(round(a2) if a2 else ""))
        ws.cell(row=i, column=kol["x =x,y"],
                value=kw.formula_xy(L["X-DXF"], L["Abmess_1"], L["Abmes_2"], i))
        ws.cell(row=i, column=kol["y =x,y"],
                value=kw.formula_xy(L["Y-DXF"], L["Abmess_1"], L["Abmes_2"], i))
        ws.cell(row=i, column=kol["UWAGI"], value=chk["uwaga"])
        ws.cell(row=i, column=kol["powod"], value=(r["powod"] if r else ""))
        barwa = raport.SEM_KOLOR.get(r["semafor"]) if r else None
        if barwa:
            ws.cell(row=i, column=kol["status"]).fill = PatternFill(
                start_color=barwa, end_color=barwa, fill_type="solid")
        sem[(r["semafor"] if r else "brak")] += 1
        if chk["uwaga"] == kw.UWAGA_OK:
            sem["UWAGI_ok"] += 1
        zap += 1

    wb.save(out_path)
    print(f"[MARKER] zaznaczono {zap} wierszy -> {out_path.name} "
          f"(kolumny od {get_column_letter(ws.max_column - len(KOL) + 1)}; oryginal nietkniety)")
    print(f"[MARKER] semafor: {dict(sem)}")
    return out_path, dict(sem)
