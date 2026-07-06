# -*- coding: utf-8 -*-
"""
RAPORT zlecenia (Etap 3-4) - scalenie ocen/raportow + zapis statusow do wykazu.

Zbiera wynik zlecenia w JEDEN obraz i (opcjonalnie) wpisuje statusy do KOPII wykazu:
  - tryb wielowariantowy (warianty.py): czyta <zeinr>_ocena.csv (zwyciezca, semafor,
    powod) + raport silnika bazowego (skala, technika, wymiar z wykazu);
  - tryb parytet (sam W-B): czyta <zeinr>_raport.csv (qc_semafor, status, skala).
Dla KAZDEJ pozycji mierzy realny wymiar wyjsciowego DXF (ezdxf.extents) i porownuje
z Abmess wykazu (bramka 1: tol = max(1 mm, 0.2%)) -> kolumna uwagi_wymiar.

technologia (G/GS/S/brak) liczona z rysunku conv.dxf gdy podany:
  spaw = tekst "Schweissgruppe"/"Welding group"; giecie = n_bend>0 z raportu
  (kontekst/wiedza/technologia-schweissgruppe-auto.md). O/GSO/SO = niewykrywalne
  z DXF (wymaga TIF/technologa) -> best-effort, operator weryfikuje sporne.

ZAPIS DO WYKAZU: zawsze do KOPII (<wykaz>_WYNIKI.xlsx). Oryginal klienta NIETKNIETY
(zasada 9 + pulapka formul: openpyxl.save kasuje cache formul NAZWA/ZAKUPY; kopie
Excel przeliczy przy otwarciu, a wymiary kontrolne bierzemy z raportu, nie z xlsx).
Kolumny dopisywane na koncu naglowka (jesli brak): status, plik_dxf, technologia,
skala, wymiar_dxf_x, wymiar_dxf_y, uwagi_wymiar, uwagi. Status barwiony semaforem.

Uzycie:
  python produkcja\\raport.py <folder_wynikow>
  python produkcja\\raport.py <folder_wynikow> <rysunek_conv.dxf>          # + technologia
  python produkcja\\raport.py <folder_wynikow> [<conv.dxf>] --wykaz <wykaz.xlsx> [--out <plik.xlsx>]

Po zmianie tutaj: python testy\\test_raport.py (PASS).
"""
import csv
import shutil
import sys
from collections import Counter
from pathlib import Path

import ezdxf
from ezdxf import bbox as _bb

# flager fazowania (kontrola/) - oznacza linie fazowania na ZOLTO + komentarz.
# Import bezpieczny: gdy modul/shapely niedostepny, raport dziala bez fazowania.
try:
    sys.path.insert(0, str(Path(__file__).resolve().parent / "kontrola"))
    import fazowanie as _fazowanie
except Exception as _e:  # pragma: no cover
    _fazowanie = None
    print(f"[RAPORT] fazowanie niedostepne ({_e}) - raport bez oznaczania fazy")

# semafor -> kolor wypelnienia (jak w Excelu: zielony/zolty/czerwony)
SEM_KOLOR = {"zielony": "C6EFCE", "zolty": "FFEB9C", "czerwony": "FFC7CE"}
SEM_IKONA = {"zielony": "🟢", "zolty": "🟡", "czerwony": "🔴"}
# kolumny dopisywane do wykazu (kolejnosc = kolejnosc dopisywania)
KOL_WYKAZ = ["status", "plik_dxf", "technologia", "skala",
             "wymiar_dxf_x", "wymiar_dxf_y", "uwagi_wymiar", "uwagi"]


def norm_semafor(s):
    """Ujednolica semafor z ocena.csv (zielony/zolty) i qc_semafor W-B (ZIELONY/...)."""
    s = (s or "").strip().lower()
    if s.startswith("ziel"):
        return "zielony"
    if s.startswith("zolt") or s.startswith("zólt") or s.startswith("żółt"):
        return "zolty"
    if s.startswith("czerw"):
        return "czerwony"
    return s or "?"


def technologia_pozycji(ma_spaw, ma_giecia):
    """Regula z 54_4867: giecia+spaw->GS, giecia->G, plaski+spaw->S, inaczej->brak.
    O/GSO/SO (obrobka) niewykrywalne z DXF - poza zakresem tej funkcji."""
    if ma_giecia and ma_spaw:
        return "GS"
    if ma_giecia:
        return "G"
    if ma_spaw:
        return "S"
    return "brak"


def has_weld(msp):
    """Rysunek ma grupe spawalnicza? Tekst 'Schweissgruppe'/'Welding group'.
    UWAGA: 'Schweissnahtvorbereitung/Welding seams' to boilerplate na KAZDYM
    rysunku - nie zawiera 'schweissgruppe', wiec substring jest bezpieczny."""
    for e in msp.query("TEXT MTEXT"):
        try:
            t = e.dxf.text if e.dxftype() == "TEXT" else e.text
        except AttributeError:
            t = ""
        t = (t or "").lower()
        if "schweissgruppe" in t or "welding group" in t:
            return True
    return False


def zmierz_dxf(path):
    """Wymiar (szerokosc, wysokosc) w mm wyjsciowego DXF; None gdy brak/pusty."""
    p = Path(path)
    if not p.exists():
        return None
    try:
        msp = ezdxf.readfile(str(p)).modelspace()
        ext = _bb.extents(msp)
        if not ext.has_data:
            return None
        return (round(ext.extmax.x - ext.extmin.x, 2),
                round(ext.extmax.y - ext.extmin.y, 2))
    except Exception:
        return None


def _wymiar_ok(dxf_wh, wykaz_pair, abs_mm=1.0, proc=0.002):
    """Bramka 1 (scisla, orientacja-agnostyczna): dopasowanie max/min z tol=max(1mm,0.2%).
    Zwraca (werdykt, szczegoly): 'ok' | 'ROZJAZD ...' | 'brak_wymiaru_wykazu' | 'brak_dxf'."""
    if not dxf_wh:
        return ("brak_dxf", "")
    if not wykaz_pair:
        return ("brak_wymiaru_wykazu", "")
    dmax, dmin = max(dxf_wh), min(dxf_wh)
    wmax, wmin = max(wykaz_pair), min(wykaz_pair)

    def tol(v):
        return max(abs_mm, proc * v)

    ok_max = abs(dmax - wmax) <= tol(wmax)
    ok_min = abs(dmin - wmin) <= tol(wmin)
    if ok_max and ok_min:
        return ("ok", "")
    szcz = []
    if not ok_max:
        szcz.append(f"max {dmax:.1f} vs {wmax:.1f}")
    if not ok_min:
        szcz.append(f"min {dmin:.1f} vs {wmin:.1f}")
    return ("ROZJAZD", "; ".join(szcz))


def _czytaj_csv(path, sep=";"):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f, delimiter=sep))


def _index_raporty(folder):
    """posn -> wiersz raportu silnika. Preferuje raport z korzenia (parytet W-B),
    potem warianty/wA, potem warianty/wB (raport bazowy wielowariantowosci)."""
    folder = Path(folder)
    kandydaci = list(folder.glob("*_raport.csv"))
    for pod in ("warianty/wA", "warianty/wB"):
        kandydaci += list((folder / pod).glob("*_raport.csv"))
    idx = {}
    for rap in kandydaci:
        try:
            for r in _czytaj_csv(rap):
                try:
                    posn = int(r["posn"])
                except (KeyError, ValueError):
                    continue
                idx.setdefault(posn, r)   # pierwszy wygrywa (korzen przed wariantami)
        except Exception as e:
            print(f"[RAPORT] nie wczytano {rap.name} ({e}) - pomijam (GLOSNO)")
    return idx


def _wykaz_pair(rap_row):
    """(Abmess_max, Abmess_min) z raportu (wykaz_w/wykaz_h); None gdy brak."""
    if not rap_row:
        return None
    try:
        w, h = float(rap_row["wykaz_w"]), float(rap_row["wykaz_h"])
        return (w, h) if w > 0 and h > 0 else None
    except (KeyError, ValueError, TypeError):
        return None


def scal(folder_wynikow, rysunek=None):
    """Scala ocene wariantow + raport silnika + realny pomiar DXF w rekordy per pozycja.
    Zwraca (zeinr, rekordy). rysunek (conv.dxf) opcjonalny: dokłada technologia."""
    folder = Path(folder_wynikow)
    if not folder.is_dir():
        raise NotADirectoryError(folder)
    raporty = _index_raporty(folder)

    ocena_csv = next(iter(sorted(folder.glob("*_ocena.csv"))), None)
    parytet_raport = next(iter(sorted(folder.glob("*_raport.csv"))), None)

    # zeinr ze zrodla ocen/raportow
    if ocena_csv:
        zeinr = ocena_csv.name.replace("_ocena.csv", "")
        baza = _czytaj_csv(ocena_csv)
        tryb = "wielowariantowy"
    elif parytet_raport:
        zeinr = parytet_raport.name.replace("_raport.csv", "")
        baza = _czytaj_csv(parytet_raport)
        tryb = "parytet"
    else:
        return (folder.name, [])

    ma_spaw = None
    if rysunek and Path(rysunek).exists():
        try:
            ma_spaw = has_weld(ezdxf.readfile(str(rysunek)).modelspace())
        except Exception as e:
            print(f"[RAPORT] technologia pominieta - nie wczytano {rysunek} ({e})")

    rekordy = []
    for b in baza:
        try:
            posn = int(b["posn"])
        except (KeyError, ValueError):
            continue
        rap = raporty.get(posn)

        if tryb == "wielowariantowy":
            semafor = norm_semafor(b.get("semafor"))
            uwagi = b.get("powod", "")
            plik_dxf = b.get("plik_produkcyjny", "-")
            zwyciezca = b.get("zwyciezca", "-")
        else:  # parytet
            semafor = norm_semafor(b.get("qc_semafor") or b.get("status"))
            uwagi = b.get("qc_powody") or b.get("status", "")
            plik_dxf = b.get("file", "-")
            zwyciezca = "W-B"

        skala = (b.get("scale") if tryb == "parytet" else None) \
            or (rap.get("scale") if rap else None) or "-"

        # technologia: giecie z raportu (n_bend), spaw z rysunku
        technologia = ""
        if ma_spaw is not None and rap is not None:
            try:
                ma_giecia = int(float(rap.get("n_bend", 0) or 0)) > 0
                technologia = technologia_pozycji(ma_spaw, ma_giecia)
            except (ValueError, TypeError):
                technologia = ""

        # realny wymiar DXF vs Abmess wykazu
        dxf_path = folder / plik_dxf if plik_dxf and plik_dxf != "-" else None
        dxf_wh = zmierz_dxf(dxf_path) if dxf_path else None
        wykaz_pair = _wykaz_pair(rap)
        werdykt_wym, szcz_wym = _wymiar_ok(dxf_wh, wykaz_pair)
        uwagi_wymiar = werdykt_wym + (f" ({szcz_wym})" if szcz_wym else "")

        # fazowanie: wykryj + oznacz linie na ZOLTO w wyjsciowym DXF + komentarz
        # (operator 2026-07-07: linie fazy ZOSTAWIC, ale kolor zolty + komentarz;
        # pozycja i tak zostaje zolta do potwierdzenia - flager nie podnosi statusu).
        if _fazowanie is not None and dxf_path and Path(dxf_path).exists():
            try:
                n_faz, faz_kom = _fazowanie.oznacz_w_pliku(dxf_path)
                if n_faz:
                    uwagi = "; ".join(x for x in (uwagi, faz_kom) if x)
                    technologia = (f"{technologia}+faza"
                                   if technologia and technologia != "brak" else "faza")
            except Exception as e:
                print(f"[RAPORT] fazowanie pominiete ({dxf_path}): {e} (GLOSNO)")

        rekordy.append(dict(
            zeinr=zeinr, posn=posn, semafor=semafor, zwyciezca=zwyciezca,
            plik_dxf=plik_dxf or "-", technologia=technologia, skala=skala,
            wymiar_dxf_x=("" if not dxf_wh else dxf_wh[0]),
            wymiar_dxf_y=("" if not dxf_wh else dxf_wh[1]),
            wykaz_x=("" if not wykaz_pair else max(wykaz_pair)),
            wykaz_y=("" if not wykaz_pair else min(wykaz_pair)),
            uwagi_wymiar=uwagi_wymiar, uwagi=uwagi))

    rekordy.sort(key=lambda r: r["posn"])
    return (zeinr, rekordy)


def zapisz_podsumowanie(folder_wynikow, zeinr, rekordy):
    """Zapisuje <zeinr>_podsumowanie.csv (deterministyczny obraz zlecenia)."""
    pola = ["zeinr", "posn", "semafor", "zwyciezca", "plik_dxf", "technologia",
            "skala", "wymiar_dxf_x", "wymiar_dxf_y", "wykaz_x", "wykaz_y",
            "uwagi_wymiar", "uwagi"]
    out = Path(folder_wynikow) / f"{zeinr}_podsumowanie.csv"
    with open(out, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=pola, delimiter=";", extrasaction="ignore")
        w.writeheader()
        w.writerows(rekordy)
    return out


def _znajdz_naglowek(ws, maks=15):
    """Wiersz naglowka = pierwszy z 'Zeinr' i 'Posn'. Zwraca (nr_wiersza, {nazwa:kol})."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=maks, values_only=True), 1):
        nazwy = {str(v).strip(): j for j, v in enumerate(row, 1) if v is not None}
        if "Zeinr" in nazwy and "Posn" in nazwy:
            return i, nazwy
    return None, {}


def zapisz_wykaz(wykaz_path, rekordy, zeinr, out_path=None):
    """Wpisuje statusy do KOPII wykazu (oryginal NIETKNIETY). Zwraca sciezke kopii
    albo None gdy sie nie udalo (GLOSNO). Kolumny KOL_WYKAZ dopisywane na koncu
    naglowka jesli brak; status barwiony semaforem."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill
    except ImportError:
        print("[RAPORT] brak openpyxl - zapis do wykazu pominiety (GLOSNO)")
        return None

    wykaz_path = Path(wykaz_path)
    if not wykaz_path.exists():
        print(f"[RAPORT] wykaz nie istnieje: {wykaz_path} (GLOSNO)")
        return None
    if out_path is None:
        out_path = wykaz_path.with_name(f"{wykaz_path.stem}_WYNIKI{wykaz_path.suffix}")
    out_path = Path(out_path)

    try:
        # kopia PRZED otwarciem - oryginal nigdy nie otwierany do zapisu
        shutil.copy(wykaz_path, out_path)
        wb = openpyxl.load_workbook(out_path)   # data_only=False: formuly jako tekst
        ws = wb.worksheets[0]
        hr, naglowek = _znajdz_naglowek(ws)
        if hr is None:
            print(f"[RAPORT] nie znaleziono naglowka Zeinr+Posn w {wykaz_path.name} (GLOSNO)")
            return None

        # kolumny docelowe: istniejace uzyj, brakujace dopisz ZA OSTATNIA UZYWANA
        # kolumna (ws.max_column+1), NIE za ostatnim NAZWANYM naglowkiem - inaczej
        # nachodza na formuly budujace NAZWA (helper cols bez naglowka; bug 2026-07-07).
        kol = dict(naglowek)
        nast = max(ws.max_column, max(naglowek.values()) if naglowek else 0) + 1
        for nazwa in KOL_WYKAZ:
            if nazwa not in kol:
                ws.cell(row=hr, column=nast, value=nazwa)
                kol[nazwa] = nast
                nast += 1

        po_posn = {int(r["posn"]): r for r in rekordy}
        zapisane = 0
        dopasowane = set()
        for i in range(hr + 1, ws.max_row + 1):
            zc = ws.cell(row=i, column=kol["Zeinr"]).value
            pc = ws.cell(row=i, column=kol["Posn"]).value
            if zc is None or zeinr not in str(zc) or pc is None:
                continue
            try:
                posn = int(pc)
            except (ValueError, TypeError):
                continue
            r = po_posn.get(posn)
            if r is None:
                continue
            dopasowane.add(posn)
            ws.cell(row=i, column=kol["status"], value=r["semafor"])
            ws.cell(row=i, column=kol["plik_dxf"], value=r["plik_dxf"])
            ws.cell(row=i, column=kol["technologia"], value=r["technologia"])
            ws.cell(row=i, column=kol["skala"], value=str(r["skala"]))
            ws.cell(row=i, column=kol["wymiar_dxf_x"], value=r["wymiar_dxf_x"])
            ws.cell(row=i, column=kol["wymiar_dxf_y"], value=r["wymiar_dxf_y"])
            ws.cell(row=i, column=kol["uwagi_wymiar"], value=r["uwagi_wymiar"])
            ws.cell(row=i, column=kol["uwagi"], value=r["uwagi"])
            barwa = SEM_KOLOR.get(r["semafor"])
            if barwa:
                ws.cell(row=i, column=kol["status"]).fill = PatternFill(
                    start_color=barwa, end_color=barwa, fill_type="solid")
            zapisane += 1

        wb.save(out_path)
        brak = sorted(set(po_posn) - dopasowane)
        if brak:
            print(f"[RAPORT] pozycje bez wiersza w wykazie (nie zapisano): {brak} (GLOSNO)")
        print(f"[RAPORT] wykaz zapisany: {out_path.name} ({zapisane} pozycji; "
              f"oryginal nietkniety)")
        return out_path
    except Exception as e:
        print(f"[RAPORT] blad zapisu do wykazu ({e}) - kopia moze byc niepelna (GLOSNO)")
        return None


def drukuj(zeinr, rekordy):
    sem = Counter(r["semafor"] for r in rekordy)
    print(f"\n=== RAPORT {zeinr}: {len(rekordy)} pozycji ===")
    for kolor in ("zielony", "zolty", "czerwony"):
        if sem.get(kolor):
            print(f"  {SEM_IKONA[kolor]} {kolor}: {sem[kolor]}")
    inne = {k: v for k, v in sem.items() if k not in SEM_IKONA}
    for k, v in inne.items():
        print(f"  ? {k}: {v}")
    ostrzez = [r for r in rekordy
               if r["semafor"] != "zielony" or not r["uwagi_wymiar"].startswith("ok")]
    if ostrzez:
        print("\ndo ogledzin (nie-zielone LUB rozjazd wymiaru):")
        for r in ostrzez:
            print(f"  {SEM_IKONA.get(r['semafor'], '?')} p{r['posn']:<3} "
                  f"{r['plik_dxf']:20} wymiar={r['uwagi_wymiar']:18} "
                  f"tech={r['technologia'] or '-':4} | {r['uwagi'][:60]}")


def main(argv):
    if not argv or argv[0] in ("-h", "--help"):
        print(__doc__)
        return 0 if argv else 2

    folder = argv[0]
    rysunek = None
    wykaz = None
    out = None
    rest = argv[1:]
    if "--wykaz" in rest:
        i = rest.index("--wykaz")
        wykaz = rest[i + 1]
        rest = rest[:i] + rest[i + 2:]
    if "--out" in rest:
        i = rest.index("--out")
        out = rest[i + 1]
        rest = rest[:i] + rest[i + 2:]
    if rest and Path(rest[0]).suffix.lower() == ".dxf":
        rysunek = rest[0]

    zeinr, rekordy = scal(folder, rysunek)
    if not rekordy:
        print(f"[RAPORT] brak ocena.csv/raport.csv w {folder}")
        return 1
    p = zapisz_podsumowanie(folder, zeinr, rekordy)
    drukuj(zeinr, rekordy)
    print(f"\npodsumowanie: {p}")
    if wykaz:
        zapisz_wykaz(wykaz, rekordy, zeinr, out)
    return 0


if __name__ == "__main__":
    try:
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass
    raise SystemExit(main(sys.argv[1:]))
