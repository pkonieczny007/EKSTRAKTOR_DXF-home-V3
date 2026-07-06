# -*- coding: utf-8 -*-
"""Testy KONTROLI WYMIARU (standard operatora): produkcja/kontrola/kontrola_wymiaru.py.

Metoda operatora (archiwum/tmp/3.2.2 - dxf_reader.py + pl.PLIK_DO_SPRAWDZENIA):
  x =x,y : OR(X==Abm1, X==Abm2, X==Abm1+-1, X==Abm2+-1); y analogicznie;
  PRAWDA i PRAWDA -> UWAGI=ok; chocby jeden FALSZ -> sprawdz.
Zweryfikowane: 14/14 zgodne z pl.PLIK_DO_SPRAWDZENIA (jesli obecny - test C).

Uzycie:  python testy\\test_kontrola_wymiaru.py     (exit 0 = PASS)
"""
import io
import sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

HERE = Path(__file__).parent
ROOT = HERE.parent
sys.path.insert(0, str(ROOT / "produkcja" / "kontrola"))
import kontrola_wymiaru as kw  # noqa: E402

FAILS = []
CHECKS = 0


def check(nazwa, war, info=""):
    global CHECKS
    CHECKS += 1
    if not war:
        FAILS.append(f"{nazwa}: {info}")


def test_metoda():
    print("--- A) metoda x=x,y / y=x,y (PRAWDA PRAWDA -> ok) ---")
    # (X, Y, Abm1, Abm2, oczekiwana_uwaga)
    P = [
        (486, 313, 471, 314, "sprawdz"),   # X 486 != 471/314 (+-1) -> falsz -> sprawdz
        (2252, 222, 2253, 223, "ok"),       # 2252=2253-1, 222=223-1 -> ok
        (441, 428, 440, 428, "ok"),         # 441=440+1, 428=428 -> ok
        (320, 63, 64, 320, "ok"),           # orientacja: X=320=Abm2, Y=63=64-1 -> ok
        (100, 50, 100, 50, "ok"),           # dokladnie
        (100, 50, 50, 100, "ok"),           # zamiana X/Y (orientacja-agnostyczna)
        (102, 50, 100, 50, "sprawdz"),      # X 102 poza +-1 -> sprawdz
        (1000, 596, 597, 1000, "ok"),       # niemiecki: Abm2=1000, X=1000, Y=596=597-1
    ]
    for x, y, a1, a2, exp in P:
        r = kw.sprawdz((x, y), a1, a2)
        check(f"A {x}x{y} vs {a1},{a2}", r["uwaga"] == exp,
              f"dostalem {r['uwaga']} (ix={r['ix']} jy={r['jy']}), oczek {exp}")
    print(f"  sprawdzono {len(P)} przypadkow")


def test_brak():
    print("\n--- B) brak DXF / brak wymiaru -> sprawdz ---")
    check("B0 brak dxf", kw.sprawdz(None, 100, 50)["uwaga"] == "sprawdz")
    check("B1 brak wymiaru", kw.sprawdz((100, 50), None, None)["uwaga"] == "sprawdz")
    check("B2 x_dxf pusty gdy brak", kw.sprawdz(None, 100, 50)["x_dxf"] == "")


def test_formula():
    print("\n--- C) formula Excela (styl operatora) ---")
    f = kw.formula_xy("CO", "CQ", "CR", 655)
    exp = "=OR(CO655=CQ655,CO655=CR655,CO655=CQ655+1,CO655=CR655+1,CO655=CQ655-1,CO655=CR655-1)"
    check("C formula", f == exp, f"\n   dostalem: {f}\n   oczek:    {exp}")
    print(f"  {f}")


def test_skala():
    print("\n--- D) skala (1:1 gdy X,Y w Abmess) ---")
    check("D0 1:1", kw.skala((100, 50), 100, 50) == "1:1")
    check("D1 skala !=1", isinstance(kw.skala((20, 10), 100, 50), float))


def test_golden_plik():
    print("\n--- E) zgodnosc z pl.PLIK_DO_SPRAWDZENIA (jesli obecny) ---")
    p = ROOT / "archiwum" / "tmp" / "pl.PLIK_DO_SPRAWDZENIA.xlsx"
    if not p.exists():
        print("  (pominieto - brak pliku referencyjnego)")
        return
    import openpyxl
    ws = openpyxl.load_workbook(p, data_only=True).active
    zg = roz = 0
    for i in range(2, ws.max_row + 1):
        x, y = ws.cell(i, 3).value, ws.cell(i, 4).value
        ich, f, g = ws.cell(i, 5).value, ws.cell(i, 6).value, ws.cell(i, 7).value
        if x is None or f is None:
            continue
        moj = kw.sprawdz((x, y), f, g)["uwaga"]
        ich_n = "ok" if (ich and str(ich).strip().lower() == "ok") else "sprawdz"
        if moj == ich_n:
            zg += 1
        else:
            roz += 1
    check("E zgodnosc z operatorem", roz == 0, f"rozjazdow {roz} / zgodnych {zg}")
    print(f"  zgodne z pl.PLIK: {zg}, rozjazdy: {roz}")


def main():
    print("=== TESTY KONTROLI WYMIARU (standard operatora) ===\n")
    test_metoda()
    test_brak()
    test_formula()
    test_skala()
    test_golden_plik()
    print(f"\nSprawdzen: {CHECKS} | Bledow: {len(FAILS)}")
    if FAILS:
        print("\n=== KONTROLA WYMIARU: FAIL ===")
        for x in FAILS:
            print(f"  - {x}")
        return 1
    print("\n=== KONTROLA WYMIARU: PASS ===")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
