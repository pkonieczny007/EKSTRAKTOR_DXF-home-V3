# -*- coding: utf-8 -*-
"""Testy MARKERA WYKAZU (Etap 4): produkcja/kontrola/marker_wykazu.py.

Zaznacza wyniki ekstrakcji w KOPII wykazu WIELORYSUNKOWEGO w stylu operatora
(pl.PLIK_DO_SPRAWDZENIA): X-DXF/Y-DXF + Abmess + formuly 'x =x,y'/'y =x,y' -> UWAGI.

A) zmierz + sprawdz przez kontrola_wymiaru (int extents, tol +-1, ok/sprawdz).
B) zaznacz - pelny przebieg na syntetycznym zleceniu 2-rysunkowym:
   - ORYGINAL NIETKNIETY (bajt w bajt),
   - kolumny dopisane za OSTATNIA UZYWANA (max_column+1), NIE za nazwanym naglowkiem
     (bug 2026-07-07: nachodzenie na formule NAZWA),
   - PER-ZEINR scoping: dwa rysunki z posn=1 -> kazdy wiersz dostaje SWOJ pomiar DXF,
   - UWAGI ok (wymiar zgodny) vs sprawdz (rozjazd),
   - formuly zapisane TAKZE jako tekst (=OR(...)),
   - status barwiony semaforem,
   - filtr wierszy (tylko 'rysuj') dziala.

Uzycie:  python testy\\test_marker_wykazu.py     (exit 0 = PASS)
"""
import io
import shutil
import sys
import tempfile
from pathlib import Path

import ezdxf

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE.parent / "produkcja" / "kontrola"))
sys.path.insert(0, str(HERE.parent / "produkcja"))
sys.path.insert(0, str(HERE.parent / "produkcja" / "silniki"))
import kontrola_wymiaru as kw   # noqa: E402
import raport                    # noqa: E402

FAILS = []
CHECKS = 0


def check(nazwa, warunek, info=""):
    global CHECKS
    CHECKS += 1
    if not warunek:
        FAILS.append(f"{nazwa}: {info}")


def _rect_dxf(path, w, h):
    """Prostokat w x h w poczatku ukladu -> extents = (w, h). Bez fazowania."""
    doc = ezdxf.new()
    doc.modelspace().add_lwpolyline(
        [(0, 0), (w, 0), (w, h), (0, h)], close=True)
    doc.saveas(path)


def _weld_dxf(path, tekst):
    doc = ezdxf.new()
    doc.modelspace().add_text(tekst).set_placement((0, 0))
    doc.saveas(path)


# ---------------------------------------------------------------- A
def test_kontrola_wymiaru():
    print("--- A) kontrola_wymiaru (int extents, tol +-1) ---")
    tmp = Path(tempfile.mkdtemp(prefix="mrk_kw_"))
    try:
        f = tmp / "r.dxf"
        _rect_dxf(f, 100.4, 50.9)          # int() -> (100, 50)
        xy = kw.zmierz_dxf_int(f)
        check("A1 int extents", xy == (100, 50), f"{xy}")
        # zgodny: X w {100,50+-1}, Y w {100,50+-1}
        r = kw.sprawdz(xy, 100.0, 50.0)
        check("A2 ok gdy zgodny", r["uwaga"] == kw.UWAGA_OK and r["ix"] and r["jy"], r)
        # rozjazd: X=100 pasuje, Y=50 nie pasuje do {200,80}
        r2 = kw.sprawdz(xy, 200.0, 80.0)
        check("A3 sprawdz gdy rozjazd", r2["uwaga"] == kw.UWAGA_SPR and not r2["jy"], r2)
        # tolerancja +-1: 51 zgodne z 50
        check("A4 tol +-1", kw.sprawdz((100, 51), 100.0, 50.0)["uwaga"] == kw.UWAGA_OK)
        # brak DXF -> sprawdz
        check("A5 brak dxf -> sprawdz", kw.sprawdz(None, 100.0, 50.0)["uwaga"] == kw.UWAGA_SPR)
        # skala 1:1 gdy oba w Abmess
        check("A6 skala 1:1", kw.skala((100, 50), 100.0, 50.0) == "1:1", kw.skala((100, 50), 100.0, 50.0))
        # formula tekstowa
        fm = kw.formula_xy("E", "G", "H", 5)
        check("A7 formula OR", fm.startswith("=OR(E5=G5,") and "E5=G5+1" in fm and "E5=H5-1" in fm, fm)
        print(f"  int={xy} ok/sprawdz/tol/skala/formula")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# ---------------------------------------------------------------- B
def _setup_dwa_rysunki(tmp):
    """Zlecenie 2-rysunkowe. Oba rysunki maja posn=1 (test scopingu per-zeinr):
       ZA=SLAA01 (100x50, giecie+spaw -> GS, wymiar OK),
       ZB=SLBB02 (wykaz 200x80, DXF 100x50 -> ROZJAZD, bez spawu/giecia -> brak)."""
    conv = tmp / "dok"
    conv.mkdir(parents=True, exist_ok=True)
    for zeinr, w, h, nb, weld, dxf_w, dxf_h in [
            ("SLAA01", 100.0, 50.0, 2, True, 100.0, 50.0),
            ("SLBB02", 200.0, 80.0, 0, False, 100.0, 50.0)]:
        # ocena.csv (jak z warianty_zlecenia)
        import csv
        with open(tmp / f"{zeinr}_ocena.csv", "w", encoding="utf-8-sig", newline="") as f:
            wr = csv.DictWriter(f, delimiter=";", fieldnames=[
                "zeinr", "posn", "zwyciezca", "semafor", "pewnosc", "zrodlo_prawda",
                "warianty", "interior", "plik_produkcyjny", "powod"])
            wr.writeheader()
            wr.writerow(dict(
                zeinr=zeinr, posn=1,
                zwyciezca="W-C", semafor="zielony" if zeinr == "SLAA01" else "zolty",
                pewnosc="wysoka", zrodlo_prawda=4, warianty="W-A;W-B;W-C",
                interior="W-C=4", plik_produkcyjny=f"{zeinr}_p1.dxf",
                powod="zgodnosc 2 metod" if zeinr == "SLAA01" else "rozbieznosc"))
        # raport silnika bazowego wA
        wA = tmp / "warianty" / "wA"
        wA.mkdir(parents=True, exist_ok=True)
        with open(wA / f"{zeinr}_raport.csv", "w", encoding="utf-8-sig", newline="") as f:
            wr = csv.DictWriter(f, delimiter=";", fieldnames=[
                "posn", "scale", "wykaz_w", "wykaz_h", "n_bend", "file"])
            wr.writeheader()
            wr.writerow(dict(posn=1, scale=1.0, wykaz_w=w, wykaz_h=h, n_bend=nb,
                             file=f"{zeinr}_p1.dxf"))
        _rect_dxf(tmp / f"{zeinr}_p1.dxf", dxf_w, dxf_h)
        if weld:
            _weld_dxf(conv / f"{zeinr}_1_conv.dxf", "Schweissgruppe X1")
        else:
            _weld_dxf(conv / f"{zeinr}_1_conv.dxf", "Rohteil")   # bez spawu
    return conv


def test_zaznacz():
    print("\n--- B) zaznacz (writeback do KOPII, per-zeinr, formuly) ---")
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl brak - test pominiety")
        return
    import marker_wykazu as mk   # noqa: E402

    tmp = Path(tempfile.mkdtemp(prefix="mrk_zaz_"))
    try:
        conv = _setup_dwa_rysunki(tmp)
        # wykaz: naglowek w 2. wierszu; kolumna NAZWA jako FORMULA w DALEKIEJ kolumnie (L=12)
        # -> test, ze marker dopisuje ZA max_column (12), nie za ostatnim nazwanym (F=6).
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Wykaz materialowy"])
        ws.cell(row=2, column=1, value="Zeinr")
        ws.cell(row=2, column=2, value="Posn")
        ws.cell(row=2, column=3, value="Abmess_1")
        ws.cell(row=2, column=4, value="Abmes_2")
        ws.cell(row=2, column=5, value="sprawdzenie")
        ws.cell(row=2, column=12, value="NAZWA")           # daleka kolumna (uzywana)
        rows = [("SLAA01", 1, "100", "50", "do rysuj", "=C3&\"_\"&D3"),
                ("SLBB02", 1, "200", "80", "do rysuj", "=C4&\"_\"&D4"),
                ("SLAA01", 2, "10", "10", "mamy", "=C5")]   # 'mamy' -> filtr odrzuca
        for r, (z, p, a1, a2, spr, naz) in enumerate(rows, start=3):
            ws.cell(row=r, column=1, value=z)
            ws.cell(row=r, column=2, value=p)
            ws.cell(row=r, column=3, value=a1)
            ws.cell(row=r, column=4, value=a2)
            ws.cell(row=r, column=5, value=spr)
            ws.cell(row=r, column=12, value=naz)           # formula w L
        wykaz = tmp / "wykaz.xlsx"
        wb.save(wykaz)
        oryg_bajty = wykaz.read_bytes()
        max_col_przed = 12

        out = tmp / "wykaz_WYNIKI.xlsx"
        res, stat = mk.zaznacz(tmp, wykaz, out, conv_dir=conv,
                               conv_pat="{z}_1_conv.dxf")
        check("B0 kopia powstala", res is not None and out.exists(), f"res={res}")
        check("B1 oryginal NIETKNIETY", wykaz.read_bytes() == oryg_bajty)

        wb2 = openpyxl.load_workbook(out)
        ws2 = wb2.worksheets[0]
        hr, nag = raport._znajdz_naglowek(ws2)
        check("B2 naglowek w 2. wierszu", hr == 2, f"hr={hr}")
        # KLUCZOWE: kolumny wynikowe dopisane ZA max_column (12), nie za F(6)
        check("B3 append za max_column",
              nag.get("status", 0) > max_col_przed, f"status kol={nag.get('status')}")
        check("B3b NAZWA nietknieta (formula w L)",
              ws2.cell(row=3, column=12).value == "=C3&\"_\"&D3",
              str(ws2.cell(row=3, column=12).value))

        # wiersz SLAA01 p1 (row 3): zielony, GS, wymiar ok
        check("B4 SLAA01 status zielony",
              ws2.cell(row=3, column=nag["status"]).value == "zielony",
              str(ws2.cell(row=3, column=nag["status"]).value))
        check("B4b SLAA01 X-DXF=100",
              ws2.cell(row=3, column=nag["X-DXF"]).value == 100,
              str(ws2.cell(row=3, column=nag["X-DXF"]).value))
        check("B4c SLAA01 technologia GS",
              ws2.cell(row=3, column=nag["technologia"]).value == "GS",
              str(ws2.cell(row=3, column=nag["technologia"]).value))
        check("B4d SLAA01 UWAGI ok",
              ws2.cell(row=3, column=nag["UWAGI"]).value == kw.UWAGA_OK,
              str(ws2.cell(row=3, column=nag["UWAGI"]).value))
        check("B4e SLAA01 fill zielony",
              (ws2.cell(row=3, column=nag["status"]).fill.start_color.rgb or "")
              .endswith(raport.SEM_KOLOR["zielony"]),
              str(ws2.cell(row=3, column=nag["status"]).fill.start_color.rgb))

        # wiersz SLBB02 p1 (row 4): PER-ZEINR scoping - dostaje SWOJ dxf 100x50,
        # wykaz 200x80 -> ROZJAZD (UWAGI sprawdz), technologia brak (bez spawu)
        check("B5 SLBB02 X-DXF=100 (swoj pomiar)",
              ws2.cell(row=4, column=nag["X-DXF"]).value == 100,
              str(ws2.cell(row=4, column=nag["X-DXF"]).value))
        check("B5b SLBB02 UWAGI sprawdz",
              ws2.cell(row=4, column=nag["UWAGI"]).value == kw.UWAGA_SPR,
              str(ws2.cell(row=4, column=nag["UWAGI"]).value))
        check("B5c SLBB02 status zolty",
              ws2.cell(row=4, column=nag["status"]).value == "zolty",
              str(ws2.cell(row=4, column=nag["status"]).value))

        # formuly zapisane jako tekst (=OR(...))
        fx = ws2.cell(row=3, column=nag["x =x,y"]).value
        check("B6 formula x tekst", isinstance(fx, str) and fx.startswith("=OR("), str(fx))

        # filtr: wiersz 'mamy' (SLAA01 p2, row 5) pominiety - status pusty
        check("B7 filtr odrzuca 'mamy'",
              ws2.cell(row=5, column=nag["status"]).value in (None, ""),
              str(ws2.cell(row=5, column=nag["status"]).value))
        print(f"  kopia={out.name}; oryginal nietkniety; append za kol {get_column_letter(nag['status'])}; "
              f"per-zeinr OK; stat={stat}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def main():
    print("=== TESTY MARKERA WYKAZU (Etap 4) ===\n")
    test_kontrola_wymiaru()
    test_zaznacz()
    print(f"\nSprawdzen: {CHECKS} | Bledow: {len(FAILS)}")
    if FAILS:
        print("\n=== MARKER WYKAZU: FAIL ===")
        for x in FAILS:
            print(f"  - {x}")
        return 1
    print("\n=== MARKER WYKAZU: PASS ===")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
