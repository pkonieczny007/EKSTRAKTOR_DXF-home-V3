# -*- coding: utf-8 -*-
"""Testy RAPORTU zlecenia (Etap 3-4): produkcja/raport.py.

A) _wymiar_ok - bramka 1 (tol=max(1mm,0.2%)): ok / ROZJAZD / granice / brak danych.
B) technologia_pozycji + has_weld - regula G/GS/S/brak (Schweissgruppe + giecie).
C) scal - merge ocena.csv + raport silnika + realny pomiar DXF:
   zielony+wymiar-ok, zolty+ROZJAZD, czerwony-bez-pliku; technologia z rysunku.
D) zapisz_wykaz - writeback do KOPII: kolumny dopisane, status barwiony,
   ORYGINAL NIETKNIETY, wiersz obcego Zeinr pominiety.

Uzycie:  python testy\\test_raport.py     (exit 0 = PASS)
"""
import csv
import io
import shutil
import sys
import tempfile
from pathlib import Path

import ezdxf

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE.parent / "produkcja"))
import raport  # noqa: E402

FAILS = []
CHECKS = 0


def check(nazwa, warunek, info=""):
    global CHECKS
    CHECKS += 1
    if not warunek:
        FAILS.append(f"{nazwa}: {info}")


def _rect_dxf(path, w, h):
    """DXF z prostokatem w x h w poczatku ukladu -> extents = (w, h)."""
    doc = ezdxf.new()
    doc.modelspace().add_lwpolyline(
        [(0, 0), (w, 0), (w, h), (0, h)], close=True)
    doc.saveas(path)


def _weld_dxf(path, tekst):
    doc = ezdxf.new()
    doc.modelspace().add_text(tekst).set_placement((0, 0))
    doc.saveas(path)


# ---------------------------------------------------------------- A
def test_wymiar_ok():
    print("--- A) _wymiar_ok (bramka 1) ---")
    check("A1 dokladny ok", raport._wymiar_ok((100.0, 50.0), (100.0, 50.0))[0] == "ok")
    # orientacja-agnostyczna: dxf (50,100) vs wykaz (100,50) = ok
    check("A2 obrocony ok", raport._wymiar_ok((50.0, 100.0), (100.0, 50.0))[0] == "ok")
    # 0.2% z 1000 = 2 mm; 1001.9 mieści się, 1003 nie
    check("A3 granica proc ok", raport._wymiar_ok((1001.9, 50.0), (1000.0, 50.0))[0] == "ok")
    check("A4 rozjazd", raport._wymiar_ok((1003.0, 50.0), (1000.0, 50.0))[0] == "ROZJAZD")
    # maly wymiar: tol=1mm (nie 0.2%); 50.9 ok, 52 nie
    check("A5 granica 1mm ok", raport._wymiar_ok((100.0, 50.9), (100.0, 50.0))[0] == "ok")
    check("A6 maly rozjazd", raport._wymiar_ok((100.0, 52.0), (100.0, 50.0))[0] == "ROZJAZD")
    check("A7 brak dxf", raport._wymiar_ok(None, (100.0, 50.0))[0] == "brak_dxf")
    check("A8 brak wykazu", raport._wymiar_ok((100.0, 50.0), None)[0] == "brak_wymiaru_wykazu")
    print("  ok/obrot/granice proc+1mm/rozjazd/brakujace")


# ---------------------------------------------------------------- B
def test_technologia():
    print("\n--- B) technologia + has_weld ---")
    check("B1 GS", raport.technologia_pozycji(True, True) == "GS")
    check("B2 G", raport.technologia_pozycji(False, True) == "G")
    check("B3 S", raport.technologia_pozycji(True, False) == "S")
    check("B4 brak", raport.technologia_pozycji(False, False) == "brak")
    tmp = Path(tempfile.mkdtemp(prefix="rap_weld_"))
    try:
        f1 = tmp / "weld.dxf"
        _weld_dxf(f1, "Schweissgruppe 12")
        check("B5 has_weld True", raport.has_weld(ezdxf.readfile(f1).modelspace()) is True)
        f2 = tmp / "plain.dxf"
        _weld_dxf(f2, "Schweissnahtvorbereitung")   # boilerplate, NIE spaw
        check("B6 boilerplate nie-spaw",
              raport.has_weld(ezdxf.readfile(f2).modelspace()) is False)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    print("  regula G/GS/S/brak + Schweissgruppe vs boilerplate")


# ---------------------------------------------------------------- C
def _setup_zlecenie(tmp, zeinr):
    """Buduje folder wynikow wielowariantowy: ocena.csv + warianty/wA/raport.csv +
    wyjsciowe DXF p1 (wymiar ok), p2 (ROZJAZD); p3 czerwony bez pliku."""
    # ocena.csv (jak z warianty_zlecenia)
    with open(tmp / f"{zeinr}_ocena.csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, delimiter=";", fieldnames=[
            "zeinr", "posn", "zwyciezca", "semafor", "pewnosc", "zrodlo_prawda",
            "warianty", "interior", "plik_produkcyjny", "powod"])
        w.writeheader()
        w.writerow(dict(zeinr=zeinr, posn=1, zwyciezca="W-C", semafor="zielony",
                        pewnosc="wysoka", zrodlo_prawda=4, warianty="W-A;W-B;W-C",
                        interior="W-C=4", plik_produkcyjny=f"{zeinr}_p1.dxf",
                        powod="zgodnosc 2 metod"))
        w.writerow(dict(zeinr=zeinr, posn=2, zwyciezca="W-B", semafor="zolty",
                        pewnosc="srednia", zrodlo_prawda=3, warianty="W-B;W-C",
                        interior="W-B=3", plik_produkcyjny=f"{zeinr}_p2.dxf",
                        powod="rozbieznosc wariantow"))
        w.writerow(dict(zeinr=zeinr, posn=3, zwyciezca="-", semafor="czerwony",
                        pewnosc="-", zrodlo_prawda=5, warianty="W-A;W-B;W-C",
                        interior="W-A=2", plik_produkcyjny="-",
                        powod="wszystkie warianty niekompletne -> czlowiek"))
    # raport silnika bazowego (skala, wymiar wykazu, giecie)
    wA = tmp / "warianty" / "wA"
    wA.mkdir(parents=True, exist_ok=True)
    with open(wA / f"{zeinr}_raport.csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, delimiter=";",
                           fieldnames=["posn", "scale", "wykaz_w", "wykaz_h", "n_bend", "file"])
        w.writeheader()
        w.writerow(dict(posn=1, scale=5.0, wykaz_w=100.0, wykaz_h=50.0, n_bend=2,
                        file=f"{zeinr}_p1.dxf"))     # giecie -> GS (ze spawem)
        w.writerow(dict(posn=2, scale=2.0, wykaz_w=200.0, wykaz_h=80.0, n_bend=0,
                        file=f"{zeinr}_p2.dxf"))     # bez giecia -> S (ze spawem)
        w.writerow(dict(posn=3, scale=1.0, wykaz_w=300.0, wykaz_h=120.0, n_bend=0,
                        file=""))
    # wyjsciowe DXF
    _rect_dxf(tmp / f"{zeinr}_p1.dxf", 100.0, 50.0)    # == wykaz -> ok
    _rect_dxf(tmp / f"{zeinr}_p2.dxf", 100.0, 50.0)    # != wykaz(200,80) -> ROZJAZD
    # p3 bez pliku
    # rysunek ze spawem
    _weld_dxf(tmp / f"{zeinr}_conv.dxf", "Schweissgruppe A1")
    return tmp / f"{zeinr}_conv.dxf"


def test_scal():
    print("\n--- C) scal (merge ocena+raport+pomiar DXF) ---")
    zeinr = "SLTEST01"
    tmp = Path(tempfile.mkdtemp(prefix="rap_scal_"))
    try:
        conv = _setup_zlecenie(tmp, zeinr)
        zz, rek = raport.scal(tmp, rysunek=conv)
        check("C0 zeinr", zz == zeinr, f"zeinr={zz}")
        check("C0 3 pozycje", len(rek) == 3, f"n={len(rek)}")
        po = {r["posn"]: r for r in rek}
        # p1 zielony, wymiar ok, GS (giecie+spaw)
        check("C1 p1 zielony", po[1]["semafor"] == "zielony", po[1]["semafor"])
        check("C1 p1 wymiar ok", po[1]["uwagi_wymiar"] == "ok", po[1]["uwagi_wymiar"])
        check("C1 p1 wymiar_dxf", po[1]["wymiar_dxf_x"] == 100.0 and po[1]["wymiar_dxf_y"] == 50.0,
              f"{po[1]['wymiar_dxf_x']}x{po[1]['wymiar_dxf_y']}")
        check("C1 p1 GS", po[1]["technologia"] == "GS", po[1]["technologia"])
        # p2 zolty, ROZJAZD (dxf 100x50 vs wykaz 200x80), S (spaw bez giecia)
        check("C2 p2 zolty", po[2]["semafor"] == "zolty", po[2]["semafor"])
        check("C2 p2 ROZJAZD", po[2]["uwagi_wymiar"].startswith("ROZJAZD"), po[2]["uwagi_wymiar"])
        check("C2 p2 S", po[2]["technologia"] == "S", po[2]["technologia"])
        # p3 czerwony, brak pliku
        check("C3 p3 czerwony", po[3]["semafor"] == "czerwony", po[3]["semafor"])
        check("C3 p3 brak dxf", po[3]["uwagi_wymiar"] == "brak_dxf", po[3]["uwagi_wymiar"])
        check("C3 p3 plik -", po[3]["plik_dxf"] == "-", po[3]["plik_dxf"])
        # podsumowanie.csv powstaje
        p = raport.zapisz_podsumowanie(tmp, zz, rek)
        check("C4 podsumowanie.csv", Path(p).exists())
        print(f"  p1={po[1]['semafor']}/{po[1]['uwagi_wymiar']}/{po[1]['technologia']}  "
              f"p2={po[2]['semafor']}/{po[2]['uwagi_wymiar']}/{po[2]['technologia']}  "
              f"p3={po[3]['semafor']}/{po[3]['uwagi_wymiar']}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# ---------------------------------------------------------------- D
def test_zapisz_wykaz():
    print("\n--- D) zapisz_wykaz (writeback do KOPII) ---")
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl brak - test pominiety")
        return
    zeinr = "SLTEST01"
    tmp = Path(tempfile.mkdtemp(prefix="rap_wykaz_"))
    try:
        # syntetyczny wykaz: naglowek nie w 1. wierszu (test szukania naglowka)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Wykaz materialowy - naglowek dekoracyjny"])
        ws.append(["Zeinr", "Posn", "ZAKUPY", "Abmess_1", "Abmes_2", "NAZWA"])
        ws.append([zeinr, 1, "BLACHA", "100", "50", "det1"])
        ws.append([zeinr, 2, "BLACHA", "200", "80", "det2"])
        ws.append([zeinr, 3, "BLACHA", "300", "120", "det3"])
        ws.append(["SLOBCY99", 1, "BLACHA", "10", "10", "obcy"])   # inny Zeinr - pominac
        wykaz = tmp / "wykaz.xlsx"
        wb.save(wykaz)
        oryg_bajty = wykaz.read_bytes()

        _setup_zlecenie(tmp, zeinr)
        _, rek = raport.scal(tmp, rysunek=None)
        out = raport.zapisz_wykaz(wykaz, rek, zeinr)
        check("D0 kopia powstala", out is not None and Path(out).exists())
        check("D1 oryginal nietkniety", wykaz.read_bytes() == oryg_bajty)

        wb2 = openpyxl.load_workbook(out)
        ws2 = wb2.worksheets[0]
        hr, nag = raport._znajdz_naglowek(ws2)
        check("D2 naglowek w 2. wierszu", hr == 2, f"hr={hr}")
        for k in raport.KOL_WYKAZ:
            check(f"D3 kolumna {k}", k in nag, f"brak {k}")
        # wiersz p1 (row hr+1=3): status zielony + fill
        c_status = ws2.cell(row=3, column=nag["status"])
        check("D4 p1 status zielony", c_status.value == "zielony", c_status.value)
        check("D4 p1 fill zielony",
              (c_status.fill.start_color.rgb or "").endswith(raport.SEM_KOLOR["zielony"]),
              str(c_status.fill.start_color.rgb))
        check("D4 p1 plik_dxf", ws2.cell(row=3, column=nag["plik_dxf"]).value == f"{zeinr}_p1.dxf")
        check("D4 p1 uwagi_wymiar ok",
              ws2.cell(row=3, column=nag["uwagi_wymiar"]).value == "ok")
        # wiersz p2 zolty
        check("D5 p2 status zolty", ws2.cell(row=4, column=nag["status"]).value == "zolty")
        # wiersz obcego Zeinr (row 6) - status pusty
        check("D6 obcy Zeinr pominiety",
              ws2.cell(row=6, column=nag["status"]).value in (None, ""),
              str(ws2.cell(row=6, column=nag["status"]).value))
        print(f"  kopia={Path(out).name}; oryginal nietkniety; 3 pozycje zapisane, obcy pominiety")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def main():
    print("=== TESTY RAPORTU (Etap 3-4) ===\n")
    test_wymiar_ok()
    test_technologia()
    test_scal()
    test_zapisz_wykaz()
    print(f"\nSprawdzen: {CHECKS} | Bledow: {len(FAILS)}")
    if FAILS:
        print("\n=== RAPORT: FAIL ===")
        for x in FAILS:
            print(f"  - {x}")
        return 1
    print("\n=== RAPORT: PASS ===")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
